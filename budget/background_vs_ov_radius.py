#!/usr/bin/env python3
"""
OV background vs OV radius for the fixed IV = 1226 mm design.

The OV sheet provides the radius evolution at IV = 1226 mm for two OV radii:
  - 1765 mm
  - 2230 mm (baseline)

We use the spreadsheet total-background points to define the "no optimisation"
curve and then apply the same thickness-optimisation idea used in
background_vs_hfe_iv_thinning.py, adapted to the OV:

  - No optimisation: constant wall thickness, so M_OV ∝ R_OV^2
  - Thickness optimisation: t_OV ∝ R_OV, so M_OV ∝ R_OV^3

Relative to the no-optimisation curve, the optimised curve therefore gains an
extra factor (R_OV / R_OV^0), with R_OV^0 = 2230 mm.
"""

from __future__ import annotations

from pathlib import Path
import sys

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from scipy.optimize import curve_fit


BUDGET_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BUDGET_DIR.parent
for _path in (str(PROJECT_ROOT), str(BUDGET_DIR)):
    if _path not in sys.path:
        sys.path.insert(0, _path)

from hfe_volume_to_iv_radius import (
    BASELINE_TARGET_MASS_KG_1,
    calibration_loss_tonnes_for_target,
    hfe_mass_tonnes,
)


CALIBRATION_LOSS_TONNES = calibration_loss_tonnes_for_target(BASELINE_TARGET_MASS_KG_1)


def hfe_mass_tonnes_from_radius_mm(radius_mm):
    radius_m = np.asarray(radius_mm, float) / 1000.0
    return hfe_mass_tonnes(radius_m, CALIBRATION_LOSS_TONNES)


XLSX_PATH = BUDGET_DIR / "Summary_bkgd_vs_hfe-shield.xlsx"
IV_RADIUS_MM = 1226.0
BASELINE_OV_RADIUS_MM = 2230.0
R_OV_GRID = np.linspace(1700.0, 2300.0, 500)
Z_BAND = 1.0
TOTAL_INTRINSIC_BACKGROUND = 0.55

FS_LABEL = 18
FS_TICK = 14
FS_LEGEND = 14


def load_ov_points(filepath: Path, iv_radius_mm: float) -> list[dict[str, float]]:
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    try:
        sheet = workbook["OV"]
        rows: list[dict[str, float]] = []
        for row_index in range(2, sheet.max_row + 1):
            iv_radius = sheet[f"A{row_index}"].value
            material = sheet[f"B{row_index}"].value
            ov_radius = sheet[f"C{row_index}"].value
            isotope = sheet[f"D{row_index}"].value
            if iv_radius != iv_radius_mm or material != "Nickel" or isotope != "Th232":
                continue

            th_eff = float(sheet[f"I{row_index}"].value)
            u_eff = float(sheet[f"I{row_index + 1}"].value)
            background = float(sheet[f"L{row_index}"].value)
            background_error = float(sheet[f"M{row_index}"].value)
            mass_kg = float(sheet[f"E{row_index}"].value)

            rows.append(
                {
                    "ov_radius_mm": float(ov_radius),
                    "mass_kg": mass_kg,
                    "hit_eff_total": th_eff + u_eff,
                    "background": background,
                    "background_error": background_error,
                }
            )
    finally:
        workbook.close()

    rows.sort(key=lambda row: row["ov_radius_mm"])
    if len(rows) < 2:
        raise ValueError(
            f"Expected at least two OV points for IV={iv_radius_mm:.0f} mm in {filepath}."
        )
    return rows


def load_summary_point(
    filepath: Path,
    component_name: str,
    radius_mm: float,
) -> tuple[float, float]:
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    try:
        sheet = workbook["Summary"]
        current_component = None
        for row_index in range(3, sheet.max_row + 1):
            component_value = sheet[f"A{row_index}"].value
            if component_value is not None:
                current_component = str(component_value).strip()
            point_radius = sheet[f"B{row_index}"].value
            if current_component != component_name or point_radius != radius_mm:
                continue
            return float(sheet[f"C{row_index}"].value), float(sheet[f"D{row_index}"].value)
    finally:
        workbook.close()

    raise ValueError(
        f"Could not find {component_name} point at radius {radius_mm:.0f} mm in Summary sheet."
    )


def _safe_sigma(errors: np.ndarray):
    errors = np.asarray(errors, float)
    valid_mask = np.isfinite(errors) & (errors > 0)
    if not np.any(valid_mask):
        return None
    min_positive = np.nanmin(errors[valid_mask])
    return np.where((~np.isfinite(errors)) | (errors <= 0), min_positive, errors)


def _confidence_band(curve, jacobian_first, jacobian_second, covariance, z_value):
    variance = (
        covariance[0, 0] * jacobian_first**2
        + 2.0 * covariance[0, 1] * jacobian_first * jacobian_second
        + covariance[1, 1] * jacobian_second**2
    )
    standard_error = np.sqrt(np.maximum(variance, 0.0))
    lower_band = np.clip(curve - z_value * standard_error, 1e-300, np.inf)
    upper_band = curve + z_value * standard_error
    return lower_band, upper_band


def fit_power_law_with_bands(
    radius_mm: np.ndarray,
    counts: np.ndarray,
    errors: np.ndarray,
    radius_grid_mm: np.ndarray,
    z_value: float,
):
    radius_mm = np.asarray(radius_mm, float)
    counts = np.asarray(counts, float)
    errors = np.asarray(errors, float)
    valid_mask = np.isfinite(radius_mm) & np.isfinite(counts) & (counts > 0)
    radius_fit = radius_mm[valid_mask]
    counts_fit = counts[valid_mask]
    errors_fit = errors[valid_mask]
    if radius_fit.size < 2:
        raise ValueError("Need at least two OV points to fit the fixed-thickness trend.")

    slope_guess, log_amplitude_guess = np.polyfit(np.log(radius_fit), np.log(counts_fit), 1)

    def model(x_values, amplitude, exponent):
        return amplitude * np.power(x_values, exponent)

    sigma = _safe_sigma(errors_fit)
    params, covariance = curve_fit(
        model,
        radius_fit,
        counts_fit,
        p0=(float(np.exp(log_amplitude_guess)), float(slope_guess)),
        sigma=sigma,
        absolute_sigma=True,
        maxfev=10000,
    )
    amplitude, exponent = params
    curve = model(radius_grid_mm, amplitude, exponent)
    jac_amplitude = np.power(radius_grid_mm, exponent)
    jac_exponent = amplitude * np.power(radius_grid_mm, exponent) * np.log(radius_grid_mm)
    lower_band, upper_band = _confidence_band(
        curve,
        jac_amplitude,
        jac_exponent,
        covariance,
        z_value,
    )
    return {
        "params": (amplitude, exponent),
        "cov": covariance,
        "curve": curve,
        "lo": lower_band,
        "hi": upper_band,
    }


def main() -> None:
    ov_rows = load_ov_points(XLSX_PATH, IV_RADIUS_MM)
    radius_data = np.array([row["ov_radius_mm"] for row in ov_rows], float)
    background_data = np.array([row["background"] for row in ov_rows], float)
    background_error_data = np.array([row["background_error"] for row in ov_rows], float)
    mass_data = np.array([row["mass_kg"] for row in ov_rows], float)
    hit_eff_data = np.array([row["hit_eff_total"] for row in ov_rows], float)

    baseline_index = int(np.argmin(np.abs(radius_data - BASELINE_OV_RADIUS_MM)))
    baseline_radius = float(radius_data[baseline_index])
    baseline_background = float(background_data[baseline_index])
    baseline_mass_kg = float(mass_data[baseline_index])

    non_baseline_index = 1 - baseline_index
    comparison_radius = float(radius_data[non_baseline_index])
    comparison_hit_eff = float(hit_eff_data[non_baseline_index])
    summary_background, summary_error = load_summary_point(XLSX_PATH, "OV", IV_RADIUS_MM)
    if not (
        np.isclose(baseline_background, summary_background, rtol=0.0, atol=1e-15)
        and np.isclose(background_error_data[baseline_index], summary_error, rtol=0.0, atol=1e-15)
    ):
        raise ValueError(
            "OV sheet and Summary sheet disagree for the IV=1226 mm OV background point."
        )

    fixed_fit = fit_power_law_with_bands(
        radius_data,
        background_data,
        background_error_data,
        R_OV_GRID,
        Z_BAND,
    )
    _, background_exponent = fixed_fit["params"]
    hit_eff_exponent = float(
        np.log(comparison_hit_eff / hit_eff_data[baseline_index])
        / np.log(comparison_radius / baseline_radius)
    )

    background_no_opt = fixed_fit["curve"]
    background_no_opt_lo = fixed_fit["lo"]
    background_no_opt_hi = fixed_fit["hi"]
    background_opt = background_no_opt * (R_OV_GRID / baseline_radius)
    background_opt_lo = background_no_opt_lo * (R_OV_GRID / baseline_radius)
    background_opt_hi = background_no_opt_hi * (R_OV_GRID / baseline_radius)

    hfe_mass_t = float(hfe_mass_tonnes_from_radius_mm(IV_RADIUS_MM))
    hfe_mass_label_t = int(np.rint(hfe_mass_t))

    print(f"IV radius fixed at {IV_RADIUS_MM:.0f} mm -> HFE mass = {hfe_mass_t:.2f} t")
    print(
        "Fixed-thickness trend from spreadsheet anchors: "
        f"B_OV(R) ∝ R^{background_exponent:.3f}"
    )
    print(f"Spreadsheet total hit-efficiency scaling: eps_OV(R) ∝ R^{hit_eff_exponent:.3f}")
    print(
        "Thickness-optimised OV curve: "
        "B_opt(R) = B_no_opt(R) * (R / R0), with R0 = 2230 mm"
    )
    print(
        f"Baseline OV mass at {baseline_radius:.0f} mm = {baseline_mass_kg:.1f} kg "
        "(spreadsheet, no optimisation)"
    )
    print(
        "Consistency check against Summary / background_vs_hfe_iv_ov: "
        f"B_OV(IV=1226 mm) = {summary_background:.6e} +/- {summary_error:.6e} [OK]"
    )

    fig, axis = plt.subplots(figsize=(10, 7))

    axis.semilogy(
        R_OV_GRID,
        background_no_opt,
        "--",
        linewidth=1.4,
        alpha=0.95,
        color="C3",
        label="OV: fixed thickness",
    )
    axis.fill_between(
        R_OV_GRID,
        background_no_opt_lo,
        background_no_opt_hi,
        color="C3",
        alpha=0.15,
        linewidth=0,
        label="_nolegend_",
    )
    axis.semilogy(
        R_OV_GRID,
        background_opt,
        "--",
        linewidth=1.4,
        alpha=0.95,
        color="C0",
        label=r"OV: optimised thickness ($t \propto R$)",
    )
    axis.fill_between(
        R_OV_GRID,
        background_opt_lo,
        background_opt_hi,
        color="C0",
        alpha=0.15,
        linewidth=0,
        label="_nolegend_",
    )

    axis.errorbar(
        radius_data,
        background_data,
        yerr=background_error_data,
        fmt="o",
        ms=5,
        color="C3",
        zorder=3,
        label="_nolegend_",
    )

    axis.plot(
        [baseline_radius],
        [TOTAL_INTRINSIC_BACKGROUND],
        marker="s",
        ms=6,
        color="0.15",
        markerfacecolor="none",
        linestyle="None",
        label="_nolegend_",
    )
    axis.annotate(
        "total intrinsic background",
        xy=(baseline_radius, TOTAL_INTRINSIC_BACKGROUND),
        xytext=(-10, -2),
        textcoords="offset points",
        ha="right",
        va="center",
        fontsize=FS_TICK - 2,
        color="0.15",
    )
    axis.axvline(
        baseline_radius,
        color="0.3",
        linestyle=(0, (4, 2)),
        linewidth=1.5,
    )
    axis.text(
        baseline_radius,
        0.03,
        "Baseline Design",
        rotation=90,
        ha="right",
        va="bottom",
        transform=axis.get_xaxis_transform(),
        fontsize=FS_TICK,
        color="0.3",
    )

    axis.text(
        0.03,
        0.03,
        f"HFE mass = {hfe_mass_label_t:d} t",
        transform=axis.transAxes,
        ha="left",
        va="bottom",
        fontsize=FS_TICK - 1,
        bbox={"boxstyle": "round,pad=0.3", "facecolor": "white", "alpha": 0.85, "edgecolor": "0.8"},
    )

    axis.set_xlabel("OV radius (mm)", fontsize=FS_LABEL)
    axis.set_ylabel("Background rate [cts/(y·2t·FWHM)]", fontsize=FS_LABEL)
    axis.set_yscale("log")
    axis.set_xlim(float(R_OV_GRID.min()), float(R_OV_GRID.max()))
    axis.set_ylim(1e-5, 1.0)
    axis.tick_params(axis="both", which="major", labelsize=FS_TICK)
    axis.grid(True, which="both", linestyle=":", alpha=0.4)
    axis.legend(fontsize=FS_LEGEND, loc="upper left")

    plt.tight_layout()
    output_dir = BUDGET_DIR / "figures"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{Path(__file__).stem}.png"
    plt.savefig(output_path, dpi=200)
    plt.close(fig)
    print(f"\nSaved to {output_path}")


if __name__ == "__main__":
    main()
