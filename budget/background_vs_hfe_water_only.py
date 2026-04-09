#!/usr/bin/env python3
"""
Water background vs HFE mass from the Summary sheet.

Fits the Water MC points with y(r) = A * exp(k r), and overlays the
semi-analytical Water models with and without dissolved Rn-222 using the same
shared HFE-mass axis as the other scans.
"""

from pathlib import Path
import sys
from functools import lru_cache
import math

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.optimize import curve_fit

# --- Config ---
XLSX_PATH = Path(__file__).resolve().parent / "Summary_bkgd_vs_hfe-shield.xlsx"
R_GRID = np.linspace(950, 1800, 600)  # mm
Z_BAND = 1.0  # 1-sigma confidence band; set 1.96 for ~95%
MU_HFE_2P5MEV_MM = 0.00674403  # HFE-7200 at 165 K (approx. LXe temperature), 1/mm

# Baseline radius for the total intrinsic-background marker; set to None to use
# max radius in data.
BASELINE_RADIUS_MM = None
TOTAL_INTRINSIC_BACKGROUND = 0.55
BASELINE_LABEL_Y = 0.45
X_AXIS_MAX_TONNES = 35.0

# Font sizes
FS_LABEL = 18
FS_TICK = 14
FS_LEGEND = 14

WATER_COMPONENTS = [
    "Water",
    "Water (semi-analytical)",
    "Water (semi-analytical with Rn)",
]
WATER_MC_LABEL = "Water"
WATER_SEMI_ANALYTICAL_NO_RN_LABEL = "Water (semi-analytical)"
WATER_SEMI_ANALYTICAL_WITH_RN_LABEL = "Water (semi-analytical with Rn)"
DISPLAY_LABELS = {
    WATER_MC_LABEL: "Water: Th/U",
    WATER_SEMI_ANALYTICAL_NO_RN_LABEL: "Water (semi-analytical): Th/U",
    WATER_SEMI_ANALYTICAL_WITH_RN_LABEL: "Water (semi-analytical): Th/U + Rn",
}
SECONDS_PER_YEAR = 86400.0 * 365.25
WATER_MODEL_RADIAL_BINS = 240
WATER_MODEL_AXIAL_BINS = 240

# Shared IV-radius -> HFE-mass conversion used in the HFE plots.
PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from hfe_volume_to_iv_radius import (
    BASELINE_TARGET_MASS_KG_1,
    calibration_loss_tonnes_for_target,
    hfe_mass_tonnes,
)

CALIBRATION_LOSS_TONNES = calibration_loss_tonnes_for_target(BASELINE_TARGET_MASS_KG_1)


def hfe_mass_tonnes_from_radius_mm(radius_mm):
    """Convert IV radius (mm) to HFE mass (tonnes)."""
    radius_m = np.asarray(radius_mm, float) / 1000.0
    return hfe_mass_tonnes(radius_m, CALIBRATION_LOSS_TONNES)


def load_data(filepath):
    """Load data from Excel file and return processed DataFrame."""
    data_frame = pd.read_excel(filepath, sheet_name="Summary", header=None)

    header_row_index = None
    for row_index, (_, row_series) in enumerate(data_frame.iterrows()):
        row_values = [str(value).strip().lower() for value in row_series.values]
        if all(
            header in row_values
            for header in ["component", "iv radius (mm)", "background", "error"]
        ):
            header_row_index = row_index
            break

    if header_row_index is None:
        raise ValueError("Header row with required columns not found.")

    data_frame = data_frame.iloc[header_row_index + 1 :, :4].copy()
    data_frame.columns = ["Component", "r_mm", "y", "e"]

    data_frame["Component"] = data_frame["Component"].ffill().astype(str).str.strip()
    for column in ["r_mm", "y", "e"]:
        data_frame[column] = pd.to_numeric(data_frame[column], errors="coerce")

    mapping = {
        "water": "Water",
        "water (theoretical)": WATER_SEMI_ANALYTICAL_NO_RN_LABEL,
        "water (theoretical with rn)": WATER_SEMI_ANALYTICAL_WITH_RN_LABEL,
    }
    data_frame["Component"] = (
        data_frame["Component"].str.lower().map(mapping).fillna(data_frame["Component"])
    )
    return data_frame.dropna(subset=["r_mm", "y"]).reset_index(drop=True)


def _load_water_plot_series():
    """Load the water MC and semi-analytical point series directly from the Water sheet."""
    workbook = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    try:
        water_sheet = workbook["Water"]

        def build_series(value_rows, value_col, error_col, radius_rows=None):
            if radius_rows is None:
                radius_rows = value_rows
            radii = np.array([float(water_sheet[f"A{row}"].value) for row in radius_rows], float)
            values = np.array(
                [float(water_sheet[f"{value_col}{row}"].value) for row in value_rows],
                float,
            )
            errors = np.array(
                [float(water_sheet[f"{error_col}{row}"].value) for row in value_rows],
                float,
            )
            return {
                "radius_mm": radii,
                "counts": values,
                "errors": errors,
            }

        series_by_component = {
            WATER_MC_LABEL: build_series((3, 5), "L", "N", radius_rows=(2, 5)),
            WATER_SEMI_ANALYTICAL_NO_RN_LABEL: build_series((8, 11, 14), "L", "N"),
            WATER_SEMI_ANALYTICAL_WITH_RN_LABEL: build_series((8, 11, 14), "K", "M"),
        }
    finally:
        workbook.close()

    return series_by_component


def _safe_sigma(errors: np.ndarray):
    """Return sigma array for curve_fit; if all zeros/NaN, return None (unweighted)."""
    errors = np.asarray(errors, float)
    valid_mask = np.isfinite(errors) & (errors > 0)
    if not np.any(valid_mask):
        return None
    min_positive = np.nanmin(errors[valid_mask])
    return np.where((~np.isfinite(errors)) | (errors <= 0), min_positive, errors)


def _confidence_band(curve, jacobian_first, jacobian_second, covariance, z_value):
    """Compute confidence bands via the delta method for two-parameter fits."""
    variance = (
        covariance[0, 0] * jacobian_first**2
        + 2.0 * covariance[0, 1] * jacobian_first * jacobian_second
        + covariance[1, 1] * jacobian_second**2
    )
    standard_error = np.sqrt(np.maximum(variance, 0.0))
    lower_band = np.clip(curve - z_value * standard_error, 1e-300, np.inf)
    upper_band = curve + z_value * standard_error
    return lower_band, upper_band


@lru_cache(maxsize=1)
def _load_water_semi_analytical_inputs():
    """Load the spreadsheet inputs used by the Water semi-analytical model."""
    workbook = load_workbook(XLSX_PATH, data_only=True, read_only=True)
    try:
        water_sheet = workbook["Water"]
        ov_sheet = workbook["OV"]

        tank_diameter_m = float(water_sheet["G21"].value)
        tank_height_m = float(water_sheet["H21"].value)
        mu_water_mm = float(water_sheet["I21"].value)
        # Keep the code-side HFE coefficient authoritative instead of the
        # spreadsheet cell, which has historically drifted across analyses.
        mu_hfe_mm = MU_HFE_2P5MEV_MM
        r_iv_ref_mm = float(water_sheet["H25"].value)
        r_ov_mm = float(water_sheet["B14"].value)
        density_water_kg_per_m3 = float(water_sheet["E18"].value)
        branch_th = float(water_sheet["B21"].value)
        branch_u = float(water_sheet["E21"].value)
        rn_activity_bq_per_kg = float(water_sheet["D57"].value)
        rn_activity_error_bq_per_kg = float(water_sheet["E57"].value)

        inputs = {
            "tank_radius_mm": 1000.0 * tank_diameter_m / 2.0,
            "tank_height_mm": 1000.0 * tank_height_m,
            "mu_water_mm": mu_water_mm,
            "mu_hfe_mm": mu_hfe_mm,
            "r_iv_ref_mm": r_iv_ref_mm,
            "r_ov_mm": r_ov_mm,
            "r_ov_ref_mm": r_ov_mm,
            "water_mass_kg": density_water_kg_per_m3
            * (
                math.pi * (tank_diameter_m / 2.0) ** 2 * tank_height_m
                - (4.0 / 3.0) * math.pi * (r_ov_mm / 1000.0) ** 3
            ),
            "generated_ov_events": float(ov_sheet["A16"].value),
            "isotopes": (
                {
                    "name": "Th232",
                    "activity_bq_per_kg": float(water_sheet["F14"].value),
                    "activity_error_bq_per_kg": float(water_sheet["G14"].value),
                    "ov_ref_efficiency": float(ov_sheet["I10"].value),
                    "branch_fraction": branch_th,
                },
                {
                    "name": "U238",
                    "activity_bq_per_kg": float(water_sheet["F15"].value),
                    "activity_error_bq_per_kg": float(water_sheet["G15"].value),
                    "ov_ref_efficiency": float(ov_sheet["I11"].value),
                    "branch_fraction": branch_u,
                },
                {
                    "name": "Rn222",
                    "activity_bq_per_kg": rn_activity_bq_per_kg,
                    "activity_error_bq_per_kg": rn_activity_error_bq_per_kg,
                    "ov_ref_efficiency": float(ov_sheet["I11"].value),
                    "branch_fraction": branch_u,
                },
            ),
        }
    finally:
        workbook.close()

    return inputs


@lru_cache(maxsize=1)
def _water_mean_factor():
    """
    Return the average water attenuation / geometry factor from the spreadsheet
    semi-analytical model.
    """
    inputs = _load_water_semi_analytical_inputs()
    tank_radius_mm = inputs["tank_radius_mm"]
    tank_height_mm = inputs["tank_height_mm"]
    r_ov_mm = inputs["r_ov_mm"]
    mu_water_mm = inputs["mu_water_mm"]

    radial_step = tank_radius_mm / WATER_MODEL_RADIAL_BINS
    axial_step = tank_height_mm / WATER_MODEL_AXIAL_BINS
    radial_centres = (np.arange(WATER_MODEL_RADIAL_BINS) + 0.5) * radial_step
    axial_centres = (
        -tank_height_mm / 2.0
        + (np.arange(WATER_MODEL_AXIAL_BINS) + 0.5) * axial_step
    )
    radius_mesh, height_mesh = np.meshgrid(radial_centres, axial_centres, indexing="xy")
    distance = np.sqrt(radius_mesh**2 + height_mesh**2)
    mask = distance >= r_ov_mm

    geometry_factor = np.zeros_like(distance)
    geometry_factor[mask] = (r_ov_mm**2) / (distance[mask] ** 2)

    attenuation = np.ones_like(distance)
    attenuation[mask] = np.exp(-mu_water_mm * (distance[mask] - r_ov_mm))

    volume_element = 2.0 * math.pi * radius_mesh * radial_step * axial_step
    integral = float(np.sum(geometry_factor * attenuation * volume_element))
    water_volume = (
        math.pi * tank_radius_mm**2 * tank_height_mm
        - (4.0 / 3.0) * math.pi * r_ov_mm**3
    )
    return integral / water_volume


def _ov_efficiency_uncertainty(efficiency, branch_fraction, n_ov, ov_ref_efficiency):
    """Spreadsheet-matched OV-to-water propagated efficiency uncertainty."""
    efficiency = np.asarray(efficiency, float)
    if ov_ref_efficiency > 0.0:
        return efficiency * np.sqrt(branch_fraction / (n_ov * ov_ref_efficiency))
    if branch_fraction <= 0.0:
        return np.zeros_like(efficiency)
    return np.full_like(efficiency, (1.14 * branch_fraction) / n_ov)


def _truncated_gaussian_mean(mass_kg, activity_bq_per_kg, activity_error_bq_per_kg, efficiency):
    """Return the spreadsheet-matched positive-truncated Gaussian mean."""
    rate_counts_per_sec = mass_kg * activity_bq_per_kg * efficiency
    sigma_rate = abs(mass_kg * efficiency * activity_error_bq_per_kg)
    if sigma_rate < 1e-15:
        return max(0.0, rate_counts_per_sec) * SECONDS_PER_YEAR

    z_value = rate_counts_per_sec / sigma_rate
    pdf_value = math.exp(-0.5 * z_value**2) / math.sqrt(2.0 * math.pi)
    cdf_value = 0.5 * (1.0 + math.erf(z_value / math.sqrt(2.0)))
    if cdf_value <= 0.0:
        return max(0.0, rate_counts_per_sec) * SECONDS_PER_YEAR
    return (rate_counts_per_sec + sigma_rate * pdf_value / cdf_value) * SECONDS_PER_YEAR


def _truncated_gaussian_spread(
    mass_kg,
    activity_bq_per_kg,
    activity_error_bq_per_kg,
    efficiency,
    efficiency_error,
):
    """Return the spreadsheet-matched positive-truncated Gaussian spread."""
    rate_counts_per_sec = mass_kg * activity_bq_per_kg * efficiency
    sigma_rate = math.hypot(
        mass_kg * efficiency * activity_error_bq_per_kg,
        mass_kg * activity_bq_per_kg * efficiency_error,
    )
    if sigma_rate < 1e-15:
        return 0.0

    z_value = rate_counts_per_sec / sigma_rate
    pdf_value = math.exp(-0.5 * z_value**2) / math.sqrt(2.0 * math.pi)
    cdf_value = 0.5 * (1.0 + math.erf(z_value / math.sqrt(2.0)))
    if cdf_value <= 0.0:
        return 0.0

    lambda_value = pdf_value / cdf_value
    inner_term = 1.0 - z_value * lambda_value - lambda_value**2
    if inner_term <= 0.0:
        return 0.0
    return sigma_rate * math.sqrt(inner_term) * SECONDS_PER_YEAR


def build_water_semi_analytical_curve_with_bands(
    radius_grid_mm=R_GRID,
    z_value=Z_BAND,
    include_rn=True,
):
    """
    Evaluate the spreadsheet Water semi-analytical model continuously in IV
    radius and return the mean curve plus a 1-sigma band.
    """
    inputs = _load_water_semi_analytical_inputs()
    radius_grid_mm = np.asarray(radius_grid_mm, float)

    screening = np.exp(
        -inputs["mu_hfe_mm"] * (radius_grid_mm - inputs["r_iv_ref_mm"])
    )
    geometry_scale = (inputs["r_ov_ref_mm"] / inputs["r_ov_mm"]) ** 2
    mean_factor = _water_mean_factor()

    total_curve = np.zeros_like(radius_grid_mm)
    total_sigma_sq = np.zeros_like(radius_grid_mm)
    for isotope in inputs["isotopes"]:
        if (not include_rn) and isotope["name"] == "Rn222":
            continue
        efficiency = isotope["ov_ref_efficiency"] * geometry_scale * screening * mean_factor
        efficiency_error = _ov_efficiency_uncertainty(
            efficiency,
            isotope["branch_fraction"],
            inputs["generated_ov_events"],
            isotope["ov_ref_efficiency"],
        )

        isotope_curve = np.array(
            [
                _truncated_gaussian_mean(
                    inputs["water_mass_kg"],
                    isotope["activity_bq_per_kg"],
                    isotope["activity_error_bq_per_kg"],
                    efficiency_value,
                )
                for efficiency_value in efficiency
            ]
        )
        isotope_sigma = np.array(
            [
                _truncated_gaussian_spread(
                    inputs["water_mass_kg"],
                    isotope["activity_bq_per_kg"],
                    isotope["activity_error_bq_per_kg"],
                    efficiency_value,
                    efficiency_error_value,
                )
                for efficiency_value, efficiency_error_value in zip(
                    efficiency, efficiency_error
                )
            ]
        )
        total_curve += isotope_curve
        total_sigma_sq += isotope_sigma**2

    total_sigma = np.sqrt(total_sigma_sq)
    lower_band = np.clip(total_curve - z_value * total_sigma, 1e-300, np.inf)
    upper_band = total_curve + z_value * total_sigma
    return {
        "curve": total_curve,
        "lo": lower_band,
        "hi": upper_band,
        "sigma": total_sigma,
        "mu_hfe_mm": inputs["mu_hfe_mm"],
        "mean_factor": mean_factor,
        "include_rn": include_rn,
    }


def fit_attenuation_with_bands(
    radius_mm,
    counts,
    errors,
    radius_grid_mm=R_GRID,
    z_value=Z_BAND,
):
    """Fit exponential attenuation model and compute confidence bands."""
    radius_mm = np.asarray(radius_mm, float)
    counts = np.asarray(counts, float)
    errors = np.asarray(errors, float)
    valid_mask = np.isfinite(radius_mm) & np.isfinite(counts) & (counts > 0)
    if valid_mask.sum() < 2:
        return None

    radius_fit = radius_mm[valid_mask]
    counts_fit = counts[valid_mask]
    errors_fit = errors[valid_mask]
    order = np.argsort(radius_fit)
    radius_fit = radius_fit[order]
    counts_fit = counts_fit[order]
    errors_fit = errors_fit[order]

    slope_guess, log_amplitude_guess = np.polyfit(radius_fit, np.log(counts_fit), 1)
    initial_params = (float(np.exp(log_amplitude_guess)), float(slope_guess))

    def model(x_values, amplitude, slope):
        return amplitude * np.exp(slope * x_values)

    sigma = _safe_sigma(errors_fit)
    params, covariance = curve_fit(
        model,
        radius_fit,
        counts_fit,
        p0=initial_params,
        sigma=sigma,
        absolute_sigma=True,
        maxfev=10000,
    )
    amplitude, slope = params
    curve = model(radius_grid_mm, amplitude, slope)

    jac_amplitude = np.exp(slope * radius_grid_mm)
    jac_slope = amplitude * radius_grid_mm * np.exp(slope * radius_grid_mm)
    lower_band, upper_band = _confidence_band(
        curve,
        jac_amplitude,
        jac_slope,
        covariance,
        z_value,
    )

    return {
        "params": (amplitude, slope),
        "cov": covariance,
        "curve": curve,
        "lo": lower_band,
        "hi": upper_band,
        "mu": -slope,
    }


def main():
    """Run the water-only analysis and plot results."""
    series_by_component = _load_water_plot_series()

    if not series_by_component:
        raise ValueError("No water data found in the Summary sheet.")

    baseline_radius = BASELINE_RADIUS_MM
    if baseline_radius is None:
        baseline_radius = max(
            float(np.max(component_series["radius_mm"]))
            for component_series in series_by_component.values()
        )
    baseline_mass = float(hfe_mass_tonnes_from_radius_mm(baseline_radius))

    default_colors = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
    color_map = {
        WATER_MC_LABEL: default_colors[0 % len(default_colors)],
        WATER_SEMI_ANALYTICAL_NO_RN_LABEL: default_colors[1 % len(default_colors)],
        WATER_SEMI_ANALYTICAL_WITH_RN_LABEL: default_colors[2 % len(default_colors)],
    }
    fit_curves = {}
    bands = {}

    if WATER_MC_LABEL in series_by_component:
        component_series = series_by_component[WATER_MC_LABEL]
        result = fit_attenuation_with_bands(
            component_series["radius_mm"],
            component_series["counts"],
            component_series["errors"],
        )
        if result is not None:
            amplitude, slope = result["params"]
            mu_mm_inv = result["mu"]
            fit_curves[WATER_MC_LABEL] = result["curve"]
            bands[WATER_MC_LABEL] = (result["lo"], result["hi"])
            print(
                f"{WATER_MC_LABEL}: mu = {mu_mm_inv:.4e} 1/mm "
                f"(A = {amplitude:.3g}, k = {slope:.4g})"
            )

    semi_analytical_results = {
        WATER_SEMI_ANALYTICAL_NO_RN_LABEL: build_water_semi_analytical_curve_with_bands(
            include_rn=False
        ),
        WATER_SEMI_ANALYTICAL_WITH_RN_LABEL: build_water_semi_analytical_curve_with_bands(
            include_rn=True
        ),
    }

    for component, result in semi_analytical_results.items():
        print(
            f"{component}: spreadsheet model with "
            f"mu_HFE = {result['mu_hfe_mm']:.4e} 1/mm, "
            f"F = {result['mean_factor']:.6e}"
        )
        if component in series_by_component:
            component_series = series_by_component[component]
            fit_result = fit_attenuation_with_bands(
                component_series["radius_mm"],
                component_series["counts"],
                component_series["errors"],
            )
            if fit_result is not None:
                print(
                    f"{component}: effective exp-fit mu = "
                    f"{fit_result['mu']:.4e} 1/mm"
                )
            model_points = np.interp(
                component_series["radius_mm"],
                R_GRID,
                result["curve"],
            )
            relative_diff = np.abs(model_points - component_series["counts"]) / np.maximum(
                component_series["counts"],
                1e-30,
            )
            print(
                f"{component}: max |model - sheet| / sheet = "
                f"{np.nanmax(relative_diff):.3%}"
            )

    _, axis = plt.subplots(figsize=(10, 7))
    mass_grid_t = hfe_mass_tonnes_from_radius_mm(R_GRID)

    if WATER_MC_LABEL in series_by_component:
        component_series = series_by_component[WATER_MC_LABEL]
        axis.errorbar(
            hfe_mass_tonnes_from_radius_mm(component_series["radius_mm"]),
            component_series["counts"],
            yerr=component_series["errors"],
            fmt="o",
            ms=4,
            label=DISPLAY_LABELS.get(WATER_MC_LABEL, WATER_MC_LABEL),
            color=color_map[WATER_MC_LABEL],
        )

    for component, curve in fit_curves.items():
        color = color_map.get(component, None)
        axis.semilogy(
            mass_grid_t,
            curve,
            "--",
            linewidth=1.2,
            alpha=0.9,
            color=color,
            label="_nolegend_",
        )
        if component in bands and component != "Water (semi-analytical)":
            lower_band, upper_band = bands[component]
            lower_band = np.clip(lower_band, 1e-300, np.inf)
            axis.fill_between(
                mass_grid_t,
                lower_band,
                upper_band,
                color=color,
                alpha=0.15,
                linewidth=0,
                label="_nolegend_",
            )

    for component in [WATER_SEMI_ANALYTICAL_NO_RN_LABEL]:
        result = semi_analytical_results[component]
        axis.semilogy(
            mass_grid_t,
            result["curve"],
            "--",
            linewidth=1.2,
            alpha=0.9,
            color=color_map[component],
            label=DISPLAY_LABELS[component],
        )
        axis.fill_between(
            mass_grid_t,
            np.clip(result["lo"], 1e-300, np.inf),
            result["hi"],
            color=color_map[component],
            alpha=0.15,
            linewidth=0,
            label="_nolegend_",
        )

    axis.plot(
        [baseline_mass],
        [TOTAL_INTRINSIC_BACKGROUND],
        marker="s",
        ms=7,
        color="0.15",
        markerfacecolor="none",
        linestyle="None",
        label="_nolegend_",
    )
    axis.annotate(
        "total intrinsic background",
        xy=(baseline_mass, TOTAL_INTRINSIC_BACKGROUND),
        xytext=(-10, -2),
        textcoords="offset points",
        ha="right",
        va="center",
        fontsize=FS_TICK - 2,
        color="0.15",
    )

    axis.set_xlabel("HFE mass (tonnes)", fontsize=FS_LABEL)
    axis.set_ylabel("Background rate [cts/(y·2t·FWHM)]", fontsize=FS_LABEL)
    axis.set_yscale("log")
    axis.set_xlim(float(mass_grid_t.min()), X_AXIS_MAX_TONNES)
    axis.set_ylim(1e-5, 1.0)
    axis.tick_params(axis="both", which="major", labelsize=FS_TICK)
    axis.grid(True, which="both", linestyle=":", alpha=0.4)

    axis.axvline(baseline_mass, color="0.3", linestyle="--", linewidth=1.2)
    axis.text(
        baseline_mass,
        BASELINE_LABEL_Y,
        "Baseline Design",
        rotation=90,
        va="bottom",
        ha="right",
        transform=axis.get_xaxis_transform(),
        fontsize=FS_TICK,
        color="0.3",
    )
    axis.legend(fontsize=FS_LEGEND)

    plt.tight_layout()
    output_dir = Path(__file__).resolve().parent / "figures"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{Path(__file__).stem}.png"
    plt.savefig(output_path)
    plt.close()


if __name__ == "__main__":
    main()
