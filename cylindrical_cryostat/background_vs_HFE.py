#!/usr/bin/env python3
"""
nEXO cylindrical cryostat background vs HFE thickness.

Plots the cylindrical IV/OV background contributions against the cylindrical
HFE thickness, while matching the styling of budget/background_vs_hfe_iv_ov.py
as closely as possible.

Antoine Amy — July 2025
"""
import math
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# ── Geometry & materials ─────────────────────────────────────────────────
R_TPC, H_TPC = 0.650, 1.300  # m, cylindrical TPC envelope
R_IV0, H_IV0, T_CYL_IV, T_CAP_IV = 1.410, 2.820, 0.010, 0.005   # m
R_OV0, H_OV0, T_CYL_OV, T_CAP_OV = 1.540, 3.072, 0.020, 0.010   # m
RHO_NI   = 8350        # kg·m⁻³
RHO_HFE  = 1730.3      # kg·m⁻³
MU_GAMMA = 0.00674403       # mm⁻¹, HFE-7200 at 165 K
SPREADSHEET_PATH = Path(__file__).resolve().parents[1] / "budget" / "Summary_bkgd_vs_hfe-shield.xlsx"
SPREADSHEET_IV_RADIUS_MM = 1691.0
SPREADSHEET_OV_RADIUS_MM = 2230.0

# ── Activities (mBq/kg) and their 1-σ errors ────────────────────────────
ACT_Th232, ACT_Th232_ERR = 2.65e-4, 1.09e-4
ACT_U238,  ACT_U238_ERR  = 0.0,     7.42e-4

# ── Configuration ───────────────────────────────────────────────────────
USE_AA_EFFICIENCIES = False  # Set to False to use BB efficiencies

# ── Plot styling (match background_vs_hfe_spreadsheet.py) ───────────────
FS_LABEL = 18
FS_TICK = 14
FS_LEGEND = 14
Z_BAND = 1.0  # 1-sigma band, matches background_vs_hfe_transition_box_only.py
TOTAL_INTRINSIC_BACKGROUND = 0.55
BASELINE_LABEL_Y = 0.25

# ── MC hit-efficiencies at reference shield (0.76 m sphere→cyl) ────────
# Spherical hit-efficiencies at 76 cm HFE
hit_sph_BB = {
    'IV': {'Th232': 9.720e-9, 'U238': 9.0e-9},
    'OV': {'Th232': 7.2e-9,   'U238': 0.0}
}

# --- my spherical hit efficiencies -------------------------------
hit_sph_AA = {
     'IV': {'Th232': 2.19234e-9, 'U238': 8.0e-10},
     'OV': {'Th232': 1.2579e-9,  'U238': 2.0e-10}
 }

# Select which efficiency dataset to use
hit_sph = hit_sph_AA if USE_AA_EFFICIENCIES else hit_sph_BB

TRANS_SCALE = 0.0033 / 0.0013        # ≃2.538

# ── Binomial σ on efficiency ─────────────────────────────────────────────
def eff_sigma(mu, n_decays, branch=1.0):
    σ = math.sqrt(mu * n_decays * branch) / n_decays
    return σ

N_DECAYS, BR_TL208 = 10_000_000_000, 0.3594

# ── TG mean helper (counts/sec → counts/yr) ─────────────────────────────
SECONDS_PER_YEAR = 365.25 * 86400.0
EPS = 1e-15

def calculate_bg_contribution(mass, activity, activity_err,
                              efficiency, efficiency_err):
    """
    Returns (mean_counts_per_year, TG_spread_per_year)
    mass [kg], activity [Bq/kg], activity_err [Bq/kg],
    efficiency [unitless], efficiency_err [unitless].
    """
    # instantaneous rate (counts/sec)
    t = mass * activity * efficiency
    # total 1-σ uncertainty in counts/sec
    u = math.hypot(mass * efficiency * activity_err,
                   mass * activity   * efficiency_err)

    if u < EPS:
        mean_yr = max(0.0, t) * SECONDS_PER_YEAR
        return mean_yr, 0.0

    z   = t / u
    pdf = math.exp(-0.5 * z*z) / math.sqrt(2.0 * math.pi)
    cdf = 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

    # truncated mean in cnt/sec
    mean_sec = t + u * (pdf / cdf) if cdf > EPS else 0.0
    mean_sec = max(0.0, mean_sec)

    # TG-spread factor
    term_b   = 1.0 + math.erf(z / math.sqrt(2.0))
    a_over_b = (math.exp(-0.5 * z*z) / term_b) if term_b > EPS else 0.0
    inner    = 1.0 - (z * a_over_b)/math.sqrt(8.0*math.pi) \
                   - (a_over_b**2)/(8.0*math.pi)
    spread_sec = u * math.sqrt(inner) if inner > 0 else 0.0

    # convert to counts/year
    mean_yr   = mean_sec   * SECONDS_PER_YEAR
    spread_yr = spread_sec * SECONDS_PER_YEAR
    return mean_yr, spread_yr


def calculate_bg_sigma_gaussian(mass, activity, activity_err,
                                efficiency, efficiency_err):
    """
    Return 1-sigma uncertainty (counts/year) from standard Gaussian propagation.
    """
    u = math.hypot(mass * efficiency * activity_err,
                   mass * activity   * efficiency_err)
    return u * SECONDS_PER_YEAR


def load_spreadsheet_baseline_points(filepath):
    """Load the default-design IV/OV total backgrounds from the workbook."""
    workbook = load_workbook(filepath, data_only=True, read_only=True)
    try:
        iv_sheet = workbook["IV"]
        ov_sheet = workbook["OV"]

        iv_point = None
        for row in range(2, iv_sheet.max_row + 1):
            iv_radius = iv_sheet[f"A{row}"].value
            material = iv_sheet[f"B{row}"].value
            isotope = iv_sheet[f"C{row}"].value
            if (
                iv_radius == SPREADSHEET_IV_RADIUS_MM
                and material == "Nickel"
                and isotope == "Th232"
            ):
                iv_point = {
                    "background": float(iv_sheet[f"K{row}"].value),
                    "error": float(iv_sheet[f"L{row}"].value),
                }
                break

        ov_point = None
        for row in range(2, ov_sheet.max_row + 1):
            iv_radius = ov_sheet[f"A{row}"].value
            material = ov_sheet[f"B{row}"].value
            ov_radius = ov_sheet[f"C{row}"].value
            isotope = ov_sheet[f"D{row}"].value
            if (
                iv_radius == SPREADSHEET_IV_RADIUS_MM
                and material == "Nickel"
                and ov_radius == SPREADSHEET_OV_RADIUS_MM
                and isotope == "Th232"
            ):
                ov_point = {
                    "background": float(ov_sheet[f"L{row}"].value),
                    "error": float(ov_sheet[f"M{row}"].value),
                }
                break
    finally:
        workbook.close()

    if iv_point is None or ov_point is None:
        raise ValueError("Could not load default-design IV/OV points from the spreadsheet.")
    return {"IV": iv_point, "OV": ov_point}


# ── Reference budgets ───────────────────────────────────────────────────
BUDGET_IV = 1.014e-3 + 1.241e-2
BUDGET_OV = 1.730e-3 + 1.994e-2

# ── Shield-thickness grid ────────────────────────────────────────────────
t0 = 0.76                     # m
t_max = 0.80
t_values = np.linspace(0.10, t_max, 100)
if not np.isclose(t_values, t0).any():
    t_values = np.sort(np.append(t_values, t0))

# ── Output arrays ────────────────────────────────────────────────────────
M_IV = np.zeros_like(t_values)
M_OV = np.zeros_like(t_values)
bg_IV,     bg_OV     = np.zeros_like(t_values), np.zeros_like(t_values)
bg_IV_err, bg_OV_err = np.zeros_like(t_values), np.zeros_like(t_values)

# ── Main loop ────────────────────────────────────────────────────────────
for i, t in enumerate(t_values):
    Δt = t0 - t
    atten = math.exp(MU_GAMMA * Δt * 1000.0)

    # masses
    R_IV, H_IV = R_IV0 - Δt, H_IV0 - 2*Δt
    V_IV_side = math.pi*((R_IV+T_CYL_IV)**2 - R_IV**2)*H_IV
    V_IV_caps = 2*math.pi*R_IV**2 * T_CAP_IV
    M_IV[i]   = (V_IV_side + V_IV_caps)*RHO_NI

    R_OV, H_OV = R_OV0 - Δt, H_OV0 - 2*Δt
    V_OV_side  = math.pi*((R_OV+T_CYL_OV)**2 - R_OV**2)*H_OV
    V_OV_caps  = 2*math.pi*R_OV**2 * T_CAP_OV
    M_OV[i]    = (V_OV_side + V_OV_caps)*RHO_NI

    # scaled & attenuated efficiencies
    μ_IV_Th = hit_sph['IV']['Th232'] * TRANS_SCALE * atten
    μ_IV_U  = hit_sph['IV']['U238']  * TRANS_SCALE * atten
    μ_OV_Th = hit_sph['OV']['Th232'] * TRANS_SCALE * atten
    μ_OV_U  = hit_sph['OV']['U238']  * TRANS_SCALE * atten

    σμ_IV_Th = eff_sigma(μ_IV_Th, N_DECAYS, BR_TL208)
    σμ_IV_U  = eff_sigma(μ_IV_U,  N_DECAYS)
    σμ_OV_Th = eff_sigma(μ_OV_Th, N_DECAYS, BR_TL208)
    σμ_OV_U  = eff_sigma(μ_OV_U,  N_DECAYS)

    # IV background + Gaussian 1-sigma band
    iv_mean_Th, _ = calculate_bg_contribution(
         M_IV[i],
         ACT_Th232*1e-3, ACT_Th232_ERR*1e-3,
         μ_IV_Th, σμ_IV_Th
    )
    iv_mean_U, _   = calculate_bg_contribution(
         M_IV[i],
         ACT_U238*1e-3, ACT_U238_ERR*1e-3,
         μ_IV_U, σμ_IV_U
    )
    iv_sigma_Th = calculate_bg_sigma_gaussian(
         M_IV[i],
         ACT_Th232*1e-3, ACT_Th232_ERR*1e-3,
         μ_IV_Th, σμ_IV_Th
    )
    iv_sigma_U  = calculate_bg_sigma_gaussian(
         M_IV[i],
         ACT_U238*1e-3, ACT_U238_ERR*1e-3,
         μ_IV_U, σμ_IV_U
    )
    bg_IV[i]     = iv_mean_Th + iv_mean_U
    bg_IV_err[i] = math.hypot(iv_sigma_Th, iv_sigma_U)

    # OV background + Gaussian 1-sigma band
    ov_mean_Th, _ = calculate_bg_contribution(
         M_OV[i],
         ACT_Th232*1e-3, ACT_Th232_ERR*1e-3,
         μ_OV_Th, σμ_OV_Th
    )
    ov_mean_U, _   = calculate_bg_contribution(
         M_OV[i],
         ACT_U238*1e-3, ACT_U238_ERR*1e-3,
         μ_OV_U, σμ_OV_U
    )
    ov_sigma_Th = calculate_bg_sigma_gaussian(
         M_OV[i],
         ACT_Th232*1e-3, ACT_Th232_ERR*1e-3,
         μ_OV_Th, σμ_OV_Th
    )
    ov_sigma_U  = calculate_bg_sigma_gaussian(
         M_OV[i],
         ACT_U238*1e-3, ACT_U238_ERR*1e-3,
         μ_OV_U, σμ_OV_U
    )
    bg_OV[i]     = ov_mean_Th + ov_mean_U
    bg_OV_err[i] = math.hypot(ov_sigma_Th, ov_sigma_U)

spreadsheet_points = load_spreadsheet_baseline_points(SPREADSHEET_PATH)
baseline_mask = np.isclose(t_values, t0)
baseline_idx = (
    int(np.where(baseline_mask)[0][0])
    if np.any(baseline_mask)
    else int(np.argmin(np.abs(t_values - t0)))
)

iv_bg_scale = spreadsheet_points["IV"]["background"] / bg_IV[baseline_idx]
iv_err_scale = spreadsheet_points["IV"]["error"] / bg_IV_err[baseline_idx]
ov_bg_scale = spreadsheet_points["OV"]["background"] / bg_OV[baseline_idx]
ov_err_scale = spreadsheet_points["OV"]["error"] / bg_OV_err[baseline_idx]

bg_IV *= iv_bg_scale
bg_IV_err *= iv_err_scale
bg_OV *= ov_bg_scale
bg_OV_err *= ov_err_scale

# ── Diagnostics ─────────────────────────────────────────────────────────
def show(idx, lbl):
    print(f"\n— {lbl} (t = {t_values[idx]*100:.1f} cm) —")
    print(f"Inner-vessel mass : {M_IV[idx]:9.1f} kg")
    print(f"Outer-vessel mass : {M_OV[idx]:9.1f} kg")
    print(f"BG IV  : {bg_IV[idx]:.3e} ± {bg_IV_err[idx]:.3e} cnt/y")
    print(f"BG OV  : {bg_OV[idx]:.3e} ± {bg_OV_err[idx]:.3e} cnt/y")

def find_crossing_point(background_values, budget_value, thickness_values):
    """
    Find where background crosses budget using linear interpolation.
    Returns the thickness where crossing occurs, or None if no crossing.
    """
    for i in range(len(background_values) - 1):
        if (background_values[i] <= budget_value and background_values[i+1] >= budget_value) or \
           (background_values[i] >= budget_value and background_values[i+1] <= budget_value):
            # Linear interpolation
            t1, t2 = thickness_values[i], thickness_values[i+1]
            bg1, bg2 = background_values[i], background_values[i+1]
            if bg2 != bg1:  # Avoid division by zero
                crossing_thickness = t1 + (t2 - t1) * (budget_value - bg1) / (bg2 - bg1)
                return crossing_thickness
    return None

print("Spreadsheet baseline normalization (1691 mm IV / 2230 mm OV):")
print(
    f"  IV -> {spreadsheet_points['IV']['background']:.6e} ± "
    f"{spreadsheet_points['IV']['error']:.6e} cnt/y"
)
print(
    f"  OV -> {spreadsheet_points['OV']['background']:.6e} ± "
    f"{spreadsheet_points['OV']['error']:.6e} cnt/y"
)

show(0,   "Thinnest shield")   # 10 cm
show(baseline_idx,  "Baseline shield")   # 76 cm

# ── Find crossing points ─────────────────────────────────────────────────
print("\n" + "="*60)
print("BUDGET CROSSING ANALYSIS")
print("="*60)

# Find IV crossing
iv_crossing = find_crossing_point(bg_IV, BUDGET_IV, t_values)
if iv_crossing is not None:
    print(f"IV background crosses budget at: {iv_crossing*100:.1f} cm HFE")
else:
    print("IV background does not cross budget in the range")

# Find OV crossing
ov_crossing = find_crossing_point(bg_OV, BUDGET_OV, t_values)
if ov_crossing is not None:
    print(f"OV background crosses budget at: {ov_crossing*100:.1f} cm HFE")
else:
    print("OV background does not cross budget in the range")

# Find the limiting thickness (maximum of the two crossings)
if iv_crossing is not None and ov_crossing is not None:
    limiting_thickness = max(iv_crossing, ov_crossing)
    print(f"\nLimiting thickness (both vessels): {limiting_thickness*100:.1f} cm HFE")
elif iv_crossing is not None:
    limiting_thickness = iv_crossing
    print(f"\nLimiting thickness (IV only): {limiting_thickness*100:.1f} cm HFE")
elif ov_crossing is not None:
    limiting_thickness = ov_crossing
    print(f"\nLimiting thickness (OV only): {limiting_thickness*100:.1f} cm HFE")
else:
    print("\nNo budget crossings found in the range")

# ── Plot ────────────────────────────────────────────────────────────────
fig, axis = plt.subplots(figsize=(10, 7))
thickness_grid_cm = t_values * 100.0
baseline_thickness_cm = float(thickness_grid_cm[baseline_idx])
default_colors = plt.rcParams["axes.prop_cycle"].by_key().get("color", [])
color_map = {
    "IV": default_colors[2 % len(default_colors)],
    "OV": default_colors[3 % len(default_colors)],
}

axis.semilogy(
    thickness_grid_cm,
    bg_IV,
    "--",
    linewidth=1.2,
    alpha=0.9,
    color=color_map["IV"],
    label="_nolegend_",
)
axis.semilogy(
    thickness_grid_cm,
    bg_OV,
    "--",
    linewidth=1.2,
    alpha=0.9,
    color=color_map["OV"],
    label="_nolegend_",
)

iv_lower = np.clip(bg_IV - Z_BAND * bg_IV_err, 1e-300, np.inf)
iv_upper = bg_IV + Z_BAND * bg_IV_err
ov_lower = np.clip(bg_OV - Z_BAND * bg_OV_err, 1e-300, np.inf)
ov_upper = bg_OV + Z_BAND * bg_OV_err
axis.fill_between(
    thickness_grid_cm,
    iv_lower,
    iv_upper,
    color=color_map["IV"],
    alpha=0.15,
    linewidth=0,
    label="_nolegend_",
)
axis.fill_between(
    thickness_grid_cm,
    ov_lower,
    ov_upper,
    color=color_map["OV"],
    alpha=0.15,
    linewidth=0,
    label="_nolegend_",
)

baseline_iv = bg_IV[baseline_idx]
baseline_iv_err = bg_IV_err[baseline_idx]
baseline_ov = bg_OV[baseline_idx]
baseline_ov_err = bg_OV_err[baseline_idx]
axis.errorbar(
    [baseline_thickness_cm],
    [baseline_iv],
    yerr=[baseline_iv_err],
    fmt="o",
    ms=5,
    color=color_map["IV"],
    label="Nickel IV: Th/U",
    zorder=3,
)
axis.errorbar(
    [baseline_thickness_cm],
    [baseline_ov],
    yerr=[baseline_ov_err],
    fmt="o",
    ms=5,
    color=color_map["OV"],
    label="Nickel OV: Th/U",
    zorder=3,
)

axis.plot(
    [baseline_thickness_cm],
    [TOTAL_INTRINSIC_BACKGROUND],
    marker="s",
    ms=7,
    color="0.15",
    markerfacecolor="none",
    linestyle="None",
    label="_nolegend_",
)
axis.annotate(
    "total intrinsic background",
    xy=(baseline_thickness_cm, TOTAL_INTRINSIC_BACKGROUND),
    xytext=(-10, -2),
    textcoords="offset points",
    ha="right",
    va="center",
    fontsize=FS_TICK - 2,
    color="0.15",
)

axis.set_xlabel("HFE thickness [cm]", fontsize=FS_LABEL)
axis.set_ylabel("Background rate [cts/(y·2t·FWHM)]", fontsize=FS_LABEL)
axis.set_yscale("log")
axis.set_ylim(1e-5, 1.0)
axis.tick_params(axis="both", which="major", labelsize=FS_TICK)
axis.grid(True, which="both", linestyle=":", alpha=0.4)
axis.set_xlim(float(thickness_grid_cm.min()), 80.0)

axis.axvline(
    baseline_thickness_cm,
    color="0.3",
    linestyle="--",
    linewidth=1.2,
)
axis.text(
    baseline_thickness_cm,
    BASELINE_LABEL_Y,
    "Baseline Design",
    rotation=90,
    va="bottom",
    ha="right",
    transform=axis.get_xaxis_transform(),
    fontsize=FS_TICK,
    color="0.3",
)

axis.legend(fontsize=FS_LEGEND)
plt.tight_layout()
output_dir = Path(__file__).resolve().parents[1] / "budget" / "figures"
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / f"{Path(__file__).stem}.png"
plt.savefig(output_path)
plt.close()
